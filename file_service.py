import os
import csv
import json
import re
import logging
from fileserver import config
import openpyxl
import utils

logger = logging.getLogger('app.file_service')


def read_file(path_to_file):
    """Read files on your server
    Args:
        path_to_file (srt): Path to file
    Returns:
        content (str): Return content of the file
    """
    try:
        with open(path_to_file, 'r') as file:
            content = file.read()
            return content
    except FileNotFoundError:
        logger.error("The user passed a non-existent file name for reading. Try again")


def create_file(length_name, extension, content, letter, digit):
    """Create files on your server with certain extension. If file have existed the
    will raise FileExistsError
    Args:
        length_name (srt): Define length of the file
        extension (str): Define the extension of of the file
        letter (bool): Define do it use the letters in the name of the file
        digit (bool): Define do it use the digits in the name of the file
    """
    try:
        file_name = utils.generate_file_name(length_name=length_name,
                                             letter=letter,
                                             digit=digit,
                                             extension=extension)
        with open(os.path.join(config.Configuration.UPLOAD_FOLDER, file_name), 'x') as file:
            file.write(content)
            logger.info(f'File was created with name {file_name}')
            return file_name
    except FileExistsError:
        logger.error("File wasn't created because you had tried"
                     " to create file with the existing name")


def delete_file(path_to_file):
    """Delete files from your server
    Args:
        path_to_file (srt): Path to file
    """
    try:
        os.remove(path_to_file)
        logger.info("File was deleted successfully")
    except FileNotFoundError:
        logger.error("File wasn't found. You had tried to delete not existing "
                     "file or you path to file is wrong")


def get_metadata_file(path_to_file):
    """Get meta data about your file.
    Args:
        path_to_file (srt): Path to file
    Returns:
        dict, which contains info about each file. Keys:
                user_id (str): user id
                size_file (int): file size in kilobytes
                last_access (datetime): datetime of last access to the file;
                modification_time (datetime): datetime of last file modification;
                creation_time (datetime): datetime of the creating file;
    """
    try:
        data = os.stat(path_to_file)
        return {'user_id': data.st_dev,
                'size_file': data.st_size,
                'last_access': utils.conversion_date_time(data.st_atime),
                'modification_time': utils.conversion_date_time(data.st_mtime),
                'creation_time': utils.conversion_date_time(data.st_ctime)}
    except FileNotFoundError:
        logger.error("File wasn't found for getting meta data")


def read_csv_file(path_to_file, delimiter=','):
    """Read csv file with delimiter specified by user.
        Args:
            path_to_file (srt): Path to file
            delimiter (str): Define sign for separating values. Defaults to ','.
        Returns:
            data (list): Contains all rows of the file
    """
    try:
        with open(path_to_file, newline='') as file:
            reader = csv.reader(file, delimiter=delimiter)
            data = [row for row in reader]
        return data
    except FileNotFoundError:
        logger.error("File not found for reading csv file")


def read_json_file(path_to_file):
    """Read json file.
            Args:
                path_to_file (srt): Path to file
            Returns:
                data (list): Contains all data from the file
        """

    try:
        with open(path_to_file) as f:
            data = json.load(f)
            return data
    except FileNotFoundError:
        logger.error("File for reading json file wasn't found")


def _get_cell_data(worksheet, x, y):
    """Get data from the cell in Excel file.
           Args:
               worksheet (obj): Active worksheet from Excel file
               x (int): Number of row
               y (int): Number of column
           Returns:
               value (str): Data from the cell or '' if cell was empty.
           """

    return str(worksheet.cell(row=x, column=y).value) \
        if worksheet.cell(row=x, column=y).value else ''


def read_excel_file(path_to_file):
    """Read excel file with extension .xlsx from the active sheet.
        Args:
            path_to_file (srt): Path to file
        Returns:
            data (list): Contains all rows of the active sheet of the excel file
        """

    try:
        wb = openpyxl.load_workbook(path_to_file)
        worksheet = wb.active
        return [[_get_cell_data(worksheet, row, column)
                 for column in range(1, worksheet.max_column + 1)]
                for row in range(1, worksheet.max_row + 1)]
    except FileNotFoundError:
        logger.error(f"Excel file wasn't found for reading with path: {path_to_file}")


RULE_ID = 3
RULE_DESCR = 4
CELL_ID = 6
AMOUNT = 7
EXPRESSION = 8
IF_NEGATIVE = 9
DIVIDER = 10
CELL_ID_EXPR = 11
AMOUNT_EXPR = 12


def _add_additional_columns(data):
    """Add two additional columns for concatenation cell id with expression and divider
    and amount with expression and divider.
           Args:
               data (List[list]): Data from excel file
           Returns:
               data (List[list]): Contains data from excel file and two additional columns
           """
    head = 0
    content = [data[head]]
    content[0].extend(['cell_id_expression', 'amount_expression'])
    for index_row, row in enumerate(data[1:], start=1):
        content.append(row)

        cell_id_divider = f'{row[CELL_ID]}/{row[DIVIDER]}'\
            if int(row[DIVIDER]) > 1 else row[CELL_ID]

        cell_id_expression = f"({cell_id_divider}{row[EXPRESSION].replace('(', '')}" \
            if row[EXPRESSION].startswith('(') else cell_id_divider + row[EXPRESSION]

        sign_value = row[AMOUNT] if int(row[AMOUNT]) > 0 else '0'

        amount_divider = f'{sign_value}/{row[DIVIDER]}' \
            if int(row[DIVIDER]) > 1 else sign_value

        amount_expression = f"({amount_divider}{row[EXPRESSION].replace('(', '')}" \
            if row[EXPRESSION].startswith('(') else amount_divider + row[EXPRESSION]

        content[index_row].append(cell_id_expression)
        content[index_row].append(amount_expression)
    logger.info(f"Function return content with length:{len(content)}")
    return content


SPLIT_FORMULA_SIGNS = re.compile('<=|==|<>|!=|>=|=<|=>|>|<|=')
EQUAL_SIGN = re.compile('==|<=|>=|!=|=<|=>')


def _split_formula(result_amount):
    """Split formulas on two parts by divider '<=|==|<>|!=|>=|=<|=>|>|<|='
           Args:
               result_amount (str):
           Returns:
               formulas (List): Two part of the formulas
           """
    return SPLIT_FORMULA_SIGNS.split(result_amount)


check_not_math_signs = re.compile(r'[^0-9=<>/*\-+()%! ]')


def _evaluate_amount(result_amount):
    """Evaluate the expressions of the formula'
           Args:
               result_amount (str):
           Returns:
               formulas (List): Results evaluation. Can be integer or bool
           """

    if check_not_math_signs.search(result_amount):
        logger.error(f"Formula '{result_amount}- contain not arithmetic symbols")
        return
    result_amount = result_amount.replace("%", "*1/100*")
    return eval(result_amount.replace("=", "==")) if not EQUAL_SIGN.search(result_amount) \
        else eval(result_amount.replace("<>", "!=").replace("=<", "<=").replace("=>", ">="))


def _combine_formulas(data):
    """Combine rows with cell_id_expression and amount_expression to get full formulas
    and expression.
           Args:
               data (List[list]): Data
           Returns:
               data (List[list]): Full formulas and expressions
           """
    content = []
    for row in data[1:]:
        if row[1] == '1':
            content.append([row[RULE_ID], row[RULE_DESCR]])
            col_res_formula_txt = row[CELL_ID_EXPR]
            col_res_formula_amt = row[AMOUNT_EXPR]
        elif row[1] != '1':
            col_res_formula_txt += row[CELL_ID_EXPR]
            col_res_formula_amt += row[AMOUNT_EXPR]
        if row[1] == row[2]:
            content[len(content) - 1].extend([col_res_formula_txt, col_res_formula_amt])

    return content


def _calculate_results(data):
    """ Calculate results formulas.
           Args:
               data (List[list]): Data
           Returns:
               data (List[list]): Returns scalable results
           """
    results = [['Rule ID', 'Rule description', 'Result formula', 'Result amount', 'Status', 'LHS', 'RHS']]

    for row in data:
        results.append(row)
        status = _evaluate_amount(row[-1])
        lhs, rhs = SPLIT_FORMULA_SIGNS.split(row[-1])
        result_lhs = _evaluate_amount(lhs)
        result_rhs = _evaluate_amount(rhs)
        results[len(results) - 1].extend([status, result_lhs, result_rhs])
    return results


def parse_rules(path_to_file):
    data = read_excel_file(path_to_file)
    data = _add_additional_columns(data)
    data = _combine_formulas(data)
    return _calculate_results(data)


def update_file_txt(path_to_file, content):
    try:
        with open(path_to_file, 'w') as file:
            file.write(content)
            logger.info(f'File {os.path.basename(path_to_file)} was updated with content\n'
                        f'{content}')
    except FileNotFoundError:
        logger.error("File for updating wasn't found")
